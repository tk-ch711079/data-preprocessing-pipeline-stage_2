"""
Excel に貼り付けた進捗情報に対して、名称埋め込み・不要列削除・列名変更を行う後処理ツール。

主な処理：
- CSV → 進捗履歴情報シートへの貼り付け
- マスタシートの辞書化
- 名称埋め込み（カテゴリ名 / 会社名 / プラント名 / 作業名）
- 不要列削除
- 列名変更
"""

import csv
import os
from openpyxl import load_workbook


class ProgressPostProcessor:
    """Excel 後処理を担当するクラス"""

    def __init__(self, target_dir: str):
        self.target_dir = os.path.abspath(target_dir)

    def run(self):
        xlsx = self._find(".xlsx")
        csv_file = self._find(".csv")

        wb = load_workbook(xlsx)
        ws_history = wb["進捗履歴情報"]
        ws_progress = wb["進捗情報"]

        # CSV 貼り付け
        self._paste_csv(ws_history, csv_file)

        # マスタ辞書化
        category_map = self._load_master(wb, "カテゴリグループ", 0, 3)
        company_map  = self._load_master(wb, "会社情報", 0, 2)

        # 名称埋め込み
        for row in ws_progress.iter_rows(min_row=2):
            row[0].value = category_map.get(str(row[0].value), "")
            row[1].value = company_map.get(str(row[1].value), "")

        # 不要列削除
        delete_cols = ["優先度", "削除フラグ", "登録日時"]
        header = [c.value for c in ws_progress[1]]
        for col in reversed(delete_cols):
            if col in header:
                idx = header.index(col) + 1
                ws_progress.delete_cols(idx)
                header.pop(idx - 1)

        wb.save(xlsx)

    def _find(self, ext):
        for f in os.listdir(self.target_dir):
            if f.endswith(ext):
                return os.path.join(self.target_dir, f)
        raise FileNotFoundError

    def _paste_csv(self, ws, csv_file):
        ws.delete_rows(1, ws.max_row)
        with open(csv_file, encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)

    def _load_master(self, wb, sheet, key_col, val_col):
        ws = wb[sheet]
        return {
            str(row[key_col]): row[val_col]
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[key_col] is not None
        }
